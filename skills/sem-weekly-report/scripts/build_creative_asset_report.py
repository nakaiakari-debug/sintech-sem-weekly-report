#!/usr/bin/env python3
"""
SHERPA週次 広告アセットレポート生成スクリプト（SINTECH様式）

Input:
  広告アセット_raw のCSV。カラム:
    (segments.date), campaign.name, ad_group.name,
    asset.text_asset.text, ad_group_ad_asset_view.field_type,
    ad_group_ad_asset_view.performance_label,
    metrics.impressions, metrics.clicks, metrics.cost_micros,
    metrics.conversions

  ※ segments.date は現状未追加。追加された場合は自動で週次フィルタする。
    未追加の場合はスナップショット扱いで全行を対象とする。

Output:
  SINTECH様式Excel（5シート: サマリ / HEADLINE / DESCRIPTION / 広告グループ別 / raw）
"""
import argparse
import csv
import os
import shutil
import zipfile
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Meiryo UI", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Meiryo UI", size=12, bold=True, color="1F4E78")
BODY_FONT = Font(name="Meiryo UI", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

LABEL_ORDER = ["BEST", "GOOD", "LEARNING", "PENDING", "LOW", "NOT_APPLICABLE", ""]


def parse_int(v):
    if v is None or v == "":
        return 0
    try:
        return int(float(str(v).replace(",", "")))
    except ValueError:
        return 0


def parse_float(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return 0.0


def parse_date(s):
    return datetime.strptime(s.strip(), "%Y-%m-%d").date()


def load_csv(path):
    with open(path, "r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def has_date_col(rows):
    return any("segments.date" in r and r.get("segments.date", "").strip() for r in rows)


def filter_period(rows, start, end):
    if not has_date_col(rows):
        return rows
    s, e = parse_date(start), parse_date(end)
    out = []
    for r in rows:
        d = r.get("segments.date", "").strip()
        if not d:
            continue
        try:
            dd = parse_date(d)
        except ValueError:
            continue
        if s <= dd <= e:
            out.append(r)
    return out


DUP_NOTE = ("※ ad_group_ad_asset_view は1インプレッションを、その広告で配信された見出し・説明文それぞれに"
            "重複計上する。以下の IMP・CT・COST・CV は延べ値であり、実績の合計ではない"
            "（実測で約11倍）。CTR・CVR・CPA は分子分母が同率で膨らむため比率としては使える。")

SOURCE_OF_TRUTH_NOTE = "実績（正）: 検索クエリ側参照"


def normalize(row):
    return {
        "date": row.get("segments.date", ""),
        "campaign": row.get("campaign.name", ""),
        "ad_group": row.get("ad_group.name", ""),
        "asset_text": row.get("asset.text_asset.text", ""),
        "field_type": row.get("ad_group_ad_asset_view.field_type", ""),
        "label": row.get("ad_group_ad_asset_view.performance_label", ""),
        "imp": parse_int(row.get("metrics.impressions")),
        "clicks": parse_int(row.get("metrics.clicks")),
        "cost": parse_float(row.get("metrics.cost_micros")) / 1_000_000.0,
        "cv": parse_float(row.get("metrics.conversions")),
    }


def japanese_period(start, end):
    s, e = parse_date(start), parse_date(end)
    return f"{s.year}年{s.month}月{s.day}日 - {e.year}年{e.month}月{e.day}日"


def write_title_and_period(ws, title, period_str):
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.cell(2, 1, period_str).font = BODY_FONT


def write_header(ws, headers, row_idx=3):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row_idx, i, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER


def col_letter(i):
    return get_column_letter(i)


def sheet_summary(wb, rows, period_str, has_date):
    ws = wb.create_sheet("サマリ")
    write_title_and_period(ws, "広告アセット サマリ（SINTECH作成）", period_str)
    if not has_date:
        ws.cell(2, 4, "※ segments.date未追加のためスナップショット扱い（前週比N/A）").font = Font(name="Meiryo UI", size=10, color="C00000")

    ws.cell(3, 1, SOURCE_OF_TRUTH_NOTE).font = Font(name="Meiryo UI", size=10, bold=True, color="C00000")

    ws.cell(4, 1, "field_type × performance_label 分布（件数）").font = TITLE_FONT
    labels = sorted({r["label"] or "(空)" for r in rows},
                    key=lambda x: LABEL_ORDER.index(x) if x in LABEL_ORDER else 99)
    fts = sorted({r["field_type"] or "(空)" for r in rows})
    write_header(ws, ["field_type"] + labels + ["合計"], row_idx=5)
    for i, ft in enumerate(fts, start=6):
        ws.cell(i, 1, ft).font = BODY_FONT
        ws.cell(i, 1).border = BORDER
        total = 0
        for j, lb in enumerate(labels, start=2):
            cnt = sum(1 for r in rows if (r["field_type"] or "(空)") == ft and (r["label"] or "(空)") == lb)
            total += cnt
            c = ws.cell(i, j, cnt)
            c.font = BODY_FONT
            c.alignment = RIGHT
            c.border = BORDER
            c.number_format = "#,##0"
        c = ws.cell(i, len(labels) + 2, total)
        c.font = BODY_FONT
        c.alignment = RIGHT
        c.border = BORDER
        c.number_format = "#,##0"

    start_row = 6 + len(fts) + 2
    ws.cell(start_row, 1, "field_type別 実績サマリ（延べ値）").font = TITLE_FONT
    ws.cell(start_row + 1, 1, DUP_NOTE).font = Font(name="Meiryo UI", size=9, color="C00000")
    start_row += 1
    write_header(ws, ["field_type", "アセット数", "延べIMP", "延べCT", "CTR", "延べCOST", "延べCV", "CVR", "CPA"], row_idx=start_row + 1)
    for i, ft in enumerate(fts, start=start_row + 2):
        subset = [r for r in rows if (r["field_type"] or "(空)") == ft]
        cnt = len(subset)
        imp = sum(r["imp"] for r in subset)
        clicks = sum(r["clicks"] for r in subset)
        cost = sum(r["cost"] for r in subset)
        cv = sum(r["cv"] for r in subset)
        ctr = clicks / imp if imp else 0
        cvr = cv / clicks if clicks else 0
        cpa = cost / cv if cv else 0
        vals = [ft, cnt, imp, clicks, ctr, cost, cv, cvr, cpa]
        fmts = [None, "#,##0", "#,##0", "#,##0", "0.00%", "¥#,##0", "#,##0.00", "0.00%", "¥#,##0"]
        for j, (v, fm) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(i, j, v)
            c.font = BODY_FONT
            c.border = BORDER
            if fm:
                c.number_format = fm
                c.alignment = RIGHT

    for col_i, w in enumerate([18] + [14] * 10, start=1):
        ws.column_dimensions[col_letter(col_i)].width = w


def sheet_field_type(wb, rows, field_type, period_str):
    sheet_name = f"{field_type}_パフォーマンス別"
    ws = wb.create_sheet(sheet_name[:31])
    write_title_and_period(ws, f"{field_type} パフォーマンス別一覧（SINTECH作成）", period_str)
    write_header(ws, ["No", "パフォーマンス", "広告グループ", "アセット", "IMP", "CT", "CTR", "COST", "CV", "CVR", "CPA"], row_idx=3)

    subset = [r for r in rows if r["field_type"] == field_type]
    subset.sort(key=lambda r: (
        LABEL_ORDER.index(r["label"]) if r["label"] in LABEL_ORDER else 99,
        -r["cv"], -r["cost"]
    ))
    for i, r in enumerate(subset, start=4):
        ctr = r["clicks"] / r["imp"] if r["imp"] else 0
        cvr = r["cv"] / r["clicks"] if r["clicks"] else 0
        cpa = r["cost"] / r["cv"] if r["cv"] else 0
        vals = [i - 3, r["label"] or "(空)", r["ad_group"], r["asset_text"],
                r["imp"], r["clicks"], ctr, r["cost"], r["cv"], cvr, cpa]
        fmts = [None, None, None, None, "#,##0", "#,##0", "0.00%", "¥#,##0", "#,##0.00", "0.00%", "¥#,##0"]
        for j, (v, fm) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(i, j, v)
            c.font = BODY_FONT
            c.border = BORDER
            if j == 4:
                c.alignment = LEFT
            if fm:
                c.number_format = fm
                c.alignment = RIGHT

    for col_i, w in enumerate([5, 14, 22, 48, 10, 10, 10, 12, 10, 10, 12], start=1):
        ws.column_dimensions[col_letter(col_i)].width = w


def sheet_ad_group(wb, rows, period_str):
    ws = wb.create_sheet("広告グループ別")
    write_title_and_period(ws, "広告グループ別集計（延べ値・SINTECH作成）", period_str)
    ws.cell(3, 1, DUP_NOTE + SOURCE_OF_TRUTH_NOTE).font = Font(name="Meiryo UI", size=9, color="C00000")
    write_header(ws, ["広告グループ", "アセット数", "延べIMP", "延べCT", "CTR", "延べCOST", "延べCV", "CVR", "CPA"], row_idx=4)
    bucket = defaultdict(lambda: {"cnt": 0, "imp": 0, "clicks": 0, "cost": 0.0, "cv": 0.0})
    for r in rows:
        b = bucket[r["ad_group"] or "(未設定)"]
        b["cnt"] += 1
        b["imp"] += r["imp"]
        b["clicks"] += r["clicks"]
        b["cost"] += r["cost"]
        b["cv"] += r["cv"]
    ranked = sorted(bucket.items(), key=lambda kv: kv[1]["cost"], reverse=True)
    for i, (ag, v) in enumerate(ranked, start=5):
        ctr = v["clicks"] / v["imp"] if v["imp"] else 0
        cvr = v["cv"] / v["clicks"] if v["clicks"] else 0
        cpa = v["cost"] / v["cv"] if v["cv"] else 0
        vals = [ag, v["cnt"], v["imp"], v["clicks"], ctr, v["cost"], v["cv"], cvr, cpa]
        fmts = [None, "#,##0", "#,##0", "#,##0", "0.00%", "¥#,##0", "#,##0.00", "0.00%", "¥#,##0"]
        for j, (val, fm) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(i, j, val)
            c.font = BODY_FONT
            c.border = BORDER
            if fm:
                c.number_format = fm
                c.alignment = RIGHT
    for col_i, w in enumerate([24, 12, 12, 10, 10, 12, 10, 10, 12], start=1):
        ws.column_dimensions[col_letter(col_i)].width = w


def sheet_raw(wb, rows, period_str):
    ws = wb.create_sheet("raw")
    write_title_and_period(ws, "広告アセット raw（SINTECH作成）", period_str)
    write_header(ws, ["date", "campaign", "ad_group", "field_type", "performance_label", "asset_text", "IMP", "CT", "COST", "CV"], row_idx=3)
    for i, r in enumerate(rows, start=4):
        vals = [r["date"], r["campaign"], r["ad_group"], r["field_type"], r["label"], r["asset_text"],
                r["imp"], r["clicks"], r["cost"], r["cv"]]
        fmts = [None, None, None, None, None, None, "#,##0", "#,##0", "¥#,##0", "#,##0.00"]
        for j, (v, fm) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(i, j, v)
            c.font = BODY_FONT
            c.border = BORDER
            if j == 6:
                c.alignment = LEFT
            if fm:
                c.number_format = fm
                c.alignment = RIGHT
    for col_i, w in enumerate([12, 22, 22, 14, 16, 48, 10, 10, 12, 10], start=1):
        ws.column_dimensions[col_letter(col_i)].width = w


def fix_xlsx_corruption(path):
    tmp = path + ".tmp"
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            zout.writestr(item, zin.read(item))
    shutil.move(tmp, path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--week-start", required=True)
    ap.add_argument("--week-end", required=True)
    ap.add_argument("--output", required=True)
    a = ap.parse_args()

    raw_rows = load_csv(a.input)
    has_date = has_date_col(raw_rows)
    filtered = filter_period(raw_rows, a.week_start, a.week_end)
    rows = [normalize(r) for r in filtered]
    period_str = japanese_period(a.week_start, a.week_end)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_summary(wb, rows, period_str, has_date)
    sheet_field_type(wb, rows, "HEADLINE", period_str)
    sheet_field_type(wb, rows, "DESCRIPTION", period_str)
    sheet_ad_group(wb, rows, period_str)
    sheet_raw(wb, rows, period_str)

    os.makedirs(os.path.dirname(a.output), exist_ok=True)
    wb.save(a.output)
    fix_xlsx_corruption(a.output)
    print(f"OK: {a.output}  ({len(rows)}行 / has_date={has_date})")


if __name__ == "__main__":
    main()
