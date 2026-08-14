#!/usr/bin/env python3
"""
SHERPA週次 検索クエリレポート生成スクリプト（SINTECH様式）

Input:
  検索クエリ_raw のCSV。カラム:
    segments.date, campaign.name, ad_group.name,
    search_term_view.search_term, segments.search_term_match_type,
    metrics.impressions, metrics.clicks, metrics.ctr,
    metrics.average_cpc, metrics.cost_micros,
    metrics.conversions, metrics.conversions_from_interactions_rate,
    metrics.cost_per_conversion

Output:
  SINTECH様式Excel（7シート: サマリ / CV獲得TOP20 / 費用発生CV0ワースト20 /
  マッチタイプ別 / 広告グループ別 / 前週比較 / raw）
"""
import argparse
import csv
import os
import shutil
import zipfile
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Meiryo UI", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Meiryo UI", size=12, bold=True, color="1F4E78")
BODY_FONT = Font(name="Meiryo UI", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Side(border_style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_date(s):
    return datetime.strptime(s.strip(), "%Y-%m-%d").date()


def parse_int(v):
    if v is None or v == "":
        return 0
    try:
        return int(float(str(v).replace(",", "")))
    except ValueError:
        return 0


def parse_float(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return 0.0


def load_csv(path):
    rows = []
    with open(path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append(r)
    return rows


def filter_period(rows, start, end):
    start_d, end_d = parse_date(start), parse_date(end)
    out = []
    for r in rows:
        d = r.get("segments.date", "").strip()
        if not d:
            continue
        try:
            dd = parse_date(d)
        except ValueError:
            continue
        if start_d <= dd <= end_d:
            out.append(r)
    return out


def normalize(row):
    return {
        "date": row.get("segments.date", ""),
        "campaign": row.get("campaign.name", ""),
        "ad_group": row.get("ad_group.name", ""),
        "search_term": row.get("search_term_view.search_term", ""),
        "match_type": row.get("segments.search_term_match_type", ""),
        "imp": parse_int(row.get("metrics.impressions")),
        "clicks": parse_int(row.get("metrics.clicks")),
        "cost": parse_float(row.get("metrics.cost_micros")) / 1_000_000.0,
        "cv": parse_float(row.get("metrics.conversions")),
    }


def agg_key(rows, keyfn):
    bucket = defaultdict(lambda: {"imp": 0, "clicks": 0, "cost": 0.0, "cv": 0.0})
    for r in rows:
        k = keyfn(r)
        b = bucket[k]
        b["imp"] += r["imp"]
        b["clicks"] += r["clicks"]
        b["cost"] += r["cost"]
        b["cv"] += r["cv"]
    return bucket


def write_title_and_period(ws, title, period_str):
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.cell(2, 1, period_str).font = BODY_FONT


def write_header(ws, headers, row_idx=3):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row_idx, i, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER


def col_letter(i):
    return get_column_letter(i)


def sheet_summary(wb, current, prev, period_str, prev_period_str):
    ws = wb.create_sheet("サマリ")
    write_title_and_period(ws, "検索語句レポート サマリ（SINTECH作成）", period_str)
    ws.cell(2, 4, f"前週: {prev_period_str}").font = BODY_FONT

    headers = ["指標", "当週", "前週", "差分", "前週比"]
    write_header(ws, headers, row_idx=4)

    def totals(rows):
        t = {"imp": 0, "clicks": 0, "cost": 0.0, "cv": 0.0}
        for r in rows:
            t["imp"] += r["imp"]
            t["clicks"] += r["clicks"]
            t["cost"] += r["cost"]
            t["cv"] += r["cv"]
        t["ctr"] = t["clicks"] / t["imp"] if t["imp"] else 0
        t["cpc"] = t["cost"] / t["clicks"] if t["clicks"] else 0
        t["cvr"] = t["cv"] / t["clicks"] if t["clicks"] else 0
        t["cpa"] = t["cost"] / t["cv"] if t["cv"] else 0
        return t

    cur = totals(current)
    pre = totals(prev)

    metrics = [
        ("IMP", cur["imp"], pre["imp"], "int"),
        ("CT", cur["clicks"], pre["clicks"], "int"),
        ("CTR", cur["ctr"], pre["ctr"], "pct"),
        ("CPC", cur["cpc"], pre["cpc"], "yen"),
        ("COST", cur["cost"], pre["cost"], "yen"),
        ("CV", cur["cv"], pre["cv"], "float"),
        ("CVR", cur["cvr"], pre["cvr"], "pct"),
        ("CPA", cur["cpa"], pre["cpa"], "yen"),
    ]

    for i, (name, c, p, fmt) in enumerate(metrics, start=5):
        ws.cell(i, 1, name).font = BODY_FONT
        cell_c = ws.cell(i, 2, round(c, 2))
        cell_p = ws.cell(i, 3, round(p, 2))
        diff = c - p
        rate = (c / p) if p else 0
        cell_d = ws.cell(i, 4, round(diff, 2))
        cell_r = ws.cell(i, 5, rate)
        for c_ in (cell_c, cell_p, cell_d):
            c_.font = BODY_FONT
            c_.alignment = RIGHT
            c_.border = BORDER
            if fmt == "yen":
                c_.number_format = "¥#,##0"
            elif fmt == "pct":
                c_.number_format = "0.00%"
            elif fmt == "int":
                c_.number_format = "#,##0"
            else:
                c_.number_format = "#,##0.00"
        cell_r.font = BODY_FONT
        cell_r.alignment = RIGHT
        cell_r.border = BORDER
        cell_r.number_format = "0.0%"
        ws.cell(i, 1).border = BORDER

    ag = agg_key(current, lambda r: r["ad_group"])
    ws.cell(len(metrics) + 7, 1, "広告グループ別 (当週)").font = TITLE_FONT
    ag_headers = ["広告グループ", "IMP", "CT", "COST", "CV", "CTR", "CVR", "CPA"]
    write_header(ws, ag_headers, row_idx=len(metrics) + 8)
    start_row = len(metrics) + 9
    ag_sorted = sorted(ag.items(), key=lambda kv: kv[1]["cost"], reverse=True)
    for i, (k, v) in enumerate(ag_sorted, start=start_row):
        ctr = v["clicks"] / v["imp"] if v["imp"] else 0
        cvr = v["cv"] / v["clicks"] if v["clicks"] else 0
        cpa = v["cost"] / v["cv"] if v["cv"] else 0
        vals = [k, v["imp"], v["clicks"], v["cost"], v["cv"], ctr, cvr, cpa]
        fmts = [None, "#,##0", "#,##0", "¥#,##0", "#,##0.00", "0.00%", "0.00%", "¥#,##0"]
        for j, (val, fm) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(i, j, val)
            cell.font = BODY_FONT
            cell.border = BORDER
            if fm:
                cell.number_format = fm
                cell.alignment = RIGHT

    for col_i, w in enumerate([16, 14, 14, 14, 14, 14, 14], start=1):
        ws.column_dimensions[col_letter(col_i)].width = w


def sheet_top20_cv(wb, rows, period_str):
    ws = wb.create_sheet("CV獲得TOP20")
    write_title_and_period(ws, "CV獲得 検索語句 TOP20（SINTECH作成）", period_str)
    headers = ["No", "検索語句", "マッチタイプ", "広告グループ", "IMP", "CT", "CTR", "CPC", "COST", "CV", "CVR", "CPA"]
    write_header(ws, headers, row_idx=3)

    agg = agg_key(rows, lambda r: (r["search_term"], r["match_type"], r["ad_group"]))
    ranked = sorted(agg.items(), key=lambda kv: (kv[1]["cv"], kv[1]["cost"]), reverse=True)[:20]
    for i, ((term, mt, ag), v) in enumerate(ranked, start=4):
        ctr = v["clicks"] / v["imp"] if v["imp"] else 0
        cpc = v["cost"] / v["clicks"] if v["clicks"] else 0
        cvr = v["cv"] / v["clicks"] if v["clicks"] else 0
        cpa = v["cost"] / v["cv"] if v["cv"] else 0
        vals = [i - 3, term, mt, ag, v["imp"], v["clicks"], ctr, cpc, v["cost"], v["cv"], cvr, cpa]
        fmts = [None, None, None, None, "#,##0", "#,##0", "0.00%", "¥#,##0", "¥#,##0", "#,##0.00", "0.00%", "¥#,##0"]
        for j, (val, fm) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(i, j, val)
            cell.font = BODY_FONT
            cell.border = BORDER
            if fm:
                cell.number_format = fm
                cell.alignment = RIGHT
    for col_i, w in enumerate([5, 28, 12, 24, 12, 10, 10, 10, 12, 10, 10, 12], start=1):
        ws.column_dimensions[col_letter(col_i)].width = w


def sheet_waste_cv0(wb, rows, period_str):
    ws = wb.create_sheet("費用発生CV0ワースト20")
    write_title_and_period(ws, "費用発生 CV0 検索語句 ワースト20（SINTECH作成）", period_str)
    headers = ["No", "検索語句", "マッチタイプ", "広告グループ", "IMP", "CT", "CTR", "CPC", "COST"]
    write_header(ws, headers, row_idx=3)

    agg = agg_key(rows, lambda r: (r["search_term"], r["match_type"], r["ad_group"]))
    filtered = [(k, v) for k, v in agg.items() if v["cv"] == 0 and v["cost"] > 0]
    ranked = sorted(filtered, key=lambda kv: kv[1]["cost"], reverse=True)[:20]
    for i, ((term, mt, ag), v) in enumerate(ranked, start=4):
        ctr = v["clicks"] / v["imp"] if v["imp"] else 0
        cpc = v["cost"] / v["clicks"] if v["clicks"] else 0
        vals = [i - 3, term, mt, ag, v["imp"], v["clicks"], ctr, cpc, v["cost"]]
        fmts = [None, None, None, None, "#,##0", "#,##0", "0.00%", "¥#,##0", "¥#,##0"]
        for j, (val, fm) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(i, j, val)
            cell.font = BODY_FONT
            cell.border = BORDER
            if fm:
                cell.number_format = fm
                cell.alignment = RIGHT
    for col_i, w in enumerate([5, 28, 12, 24, 12, 10, 10, 10, 12], start=1):
        ws.column_dimensions[col_letter(col_i)].width = w


def sheet_match_type(wb, rows, period_str):
    ws = wb.create_sheet("マッチタイプ別")
    write_title_and_period(ws, "マッチタイプ別集計（SINTECH作成）", period_str)
    headers = ["マッチタイプ", "IMP", "CT", "CTR", "CPC", "COST", "CV", "CVR", "CPA"]
    write_header(ws, headers, row_idx=3)
    agg = agg_key(rows, lambda r: r["match_type"] or "(未設定)")
    ranked = sorted(agg.items(), key=lambda kv: kv[1]["cost"], reverse=True)
    for i, (mt, v) in enumerate(ranked, start=4):
        ctr = v["clicks"] / v["imp"] if v["imp"] else 0
        cpc = v["cost"] / v["clicks"] if v["clicks"] else 0
        cvr = v["cv"] / v["clicks"] if v["clicks"] else 0
        cpa = v["cost"] / v["cv"] if v["cv"] else 0
        vals = [mt, v["imp"], v["clicks"], ctr, cpc, v["cost"], v["cv"], cvr, cpa]
        fmts = [None, "#,##0", "#,##0", "0.00%", "¥#,##0", "¥#,##0", "#,##0.00", "0.00%", "¥#,##0"]
        for j, (val, fm) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(i, j, val)
            cell.font = BODY_FONT
            cell.border = BORDER
            if fm:
                cell.number_format = fm
                cell.alignment = RIGHT
    for col_i, w in enumerate([16, 12, 10, 10, 10, 12, 10, 10, 12], start=1):
        ws.column_dimensions[col_letter(col_i)].width = w


def sheet_ad_group(wb, rows, period_str):
    ws = wb.create_sheet("広告グループ別")
    write_title_and_period(ws, "広告グループ別集計（SINTECH作成）", period_str)
    headers = ["広告グループ", "IMP", "CT", "CTR", "CPC", "COST", "CV", "CVR", "CPA"]
    write_header(ws, headers, row_idx=3)
    agg = agg_key(rows, lambda r: r["ad_group"] or "(未設定)")
    ranked = sorted(agg.items(), key=lambda kv: kv[1]["cost"], reverse=True)
    for i, (ag, v) in enumerate(ranked, start=4):
        ctr = v["clicks"] / v["imp"] if v["imp"] else 0
        cpc = v["cost"] / v["clicks"] if v["clicks"] else 0
        cvr = v["cv"] / v["clicks"] if v["clicks"] else 0
        cpa = v["cost"] / v["cv"] if v["cv"] else 0
        vals = [ag, v["imp"], v["clicks"], ctr, cpc, v["cost"], v["cv"], cvr, cpa]
        fmts = [None, "#,##0", "#,##0", "0.00%", "¥#,##0", "¥#,##0", "#,##0.00", "0.00%", "¥#,##0"]
        for j, (val, fm) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(i, j, val)
            cell.font = BODY_FONT
            cell.border = BORDER
            if fm:
                cell.number_format = fm
                cell.alignment = RIGHT
    for col_i, w in enumerate([24, 12, 10, 10, 10, 12, 10, 10, 12], start=1):
        ws.column_dimensions[col_letter(col_i)].width = w


def sheet_week_diff(wb, current, prev, period_str, prev_period_str):
    ws = wb.create_sheet("前週比較")
    write_title_and_period(ws, "前週比較 検索語句レベル（SINTECH作成）", f"{period_str}  vs  {prev_period_str}")
    headers = ["検索語句", "マッチタイプ", "当週IMP", "前週IMP", "IMP差", "当週CT", "前週CT", "CT差", "当週COST", "前週COST", "COST差", "当週CV", "前週CV", "CV差"]
    write_header(ws, headers, row_idx=3)

    cur = agg_key(current, lambda r: (r["search_term"], r["match_type"]))
    pre = agg_key(prev, lambda r: (r["search_term"], r["match_type"]))
    keys = set(cur.keys()) | set(pre.keys())

    def diff_row(k):
        c = cur.get(k, {"imp": 0, "clicks": 0, "cost": 0.0, "cv": 0.0})
        p = pre.get(k, {"imp": 0, "clicks": 0, "cost": 0.0, "cv": 0.0})
        return (k, c, p)

    ranked = sorted((diff_row(k) for k in keys), key=lambda t: (t[1]["cost"] + t[2]["cost"]), reverse=True)
    for i, ((term, mt), c, p) in enumerate(ranked, start=4):
        vals = [term, mt, c["imp"], p["imp"], c["imp"] - p["imp"], c["clicks"], p["clicks"], c["clicks"] - p["clicks"],
                c["cost"], p["cost"], c["cost"] - p["cost"], c["cv"], p["cv"], c["cv"] - p["cv"]]
        fmts = [None, None, "#,##0", "#,##0", "#,##0", "#,##0", "#,##0", "#,##0",
                "¥#,##0", "¥#,##0", "¥#,##0", "#,##0.00", "#,##0.00", "#,##0.00"]
        for j, (val, fm) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(i, j, val)
            cell.font = BODY_FONT
            cell.border = BORDER
            if fm:
                cell.number_format = fm
                cell.alignment = RIGHT
    for col_i, w in enumerate([28, 12, 10, 10, 10, 10, 10, 10, 12, 12, 12, 10, 10, 10], start=1):
        ws.column_dimensions[col_letter(col_i)].width = w


def sheet_raw(wb, rows, period_str):
    ws = wb.create_sheet("raw")
    write_title_and_period(ws, "検索クエリ raw（期間フィルタ済み・SINTECH作成）", period_str)
    headers = ["date", "campaign", "ad_group", "search_term", "match_type", "IMP", "CT", "COST", "CV"]
    write_header(ws, headers, row_idx=3)
    for i, r in enumerate(rows, start=4):
        vals = [r["date"], r["campaign"], r["ad_group"], r["search_term"], r["match_type"],
                r["imp"], r["clicks"], r["cost"], r["cv"]]
        fmts = [None, None, None, None, None, "#,##0", "#,##0", "¥#,##0", "#,##0.00"]
        for j, (val, fm) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(i, j, val)
            cell.font = BODY_FONT
            cell.border = BORDER
            if fm:
                cell.number_format = fm
                cell.alignment = RIGHT
    for col_i, w in enumerate([12, 24, 22, 28, 12, 10, 10, 12, 10], start=1):
        ws.column_dimensions[col_letter(col_i)].width = w


def fix_xlsx_corruption(path):
    tmp = path + ".tmp"
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            zout.writestr(item, zin.read(item))
    shutil.move(tmp, path)


def japanese_period(start, end):
    s, e = parse_date(start), parse_date(end)
    return f"{s.year}年{s.month}月{s.day}日 - {e.year}年{e.month}月{e.day}日"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--week-start", required=True)
    ap.add_argument("--week-end", required=True)
    ap.add_argument("--prev-start", required=True)
    ap.add_argument("--prev-end", required=True)
    ap.add_argument("--output", required=True)
    a = ap.parse_args()

    rows_all = load_csv(a.input)
    cur_rows = [normalize(r) for r in filter_period(rows_all, a.week_start, a.week_end)]
    pre_rows = [normalize(r) for r in filter_period(rows_all, a.prev_start, a.prev_end)]

    period_str = japanese_period(a.week_start, a.week_end)
    prev_period_str = japanese_period(a.prev_start, a.prev_end)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_summary(wb, cur_rows, pre_rows, period_str, prev_period_str)
    sheet_top20_cv(wb, cur_rows, period_str)
    sheet_waste_cv0(wb, cur_rows, period_str)
    sheet_match_type(wb, cur_rows, period_str)
    sheet_ad_group(wb, cur_rows, period_str)
    sheet_week_diff(wb, cur_rows, pre_rows, period_str, prev_period_str)
    sheet_raw(wb, cur_rows, period_str)

    os.makedirs(os.path.dirname(a.output), exist_ok=True)
    wb.save(a.output)
    fix_xlsx_corruption(a.output)
    print(f"OK: {a.output}  (当週 {len(cur_rows)}行 / 前週 {len(pre_rows)}行)")


if __name__ == "__main__":
    main()
